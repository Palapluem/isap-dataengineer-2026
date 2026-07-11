from __future__ import annotations

from pathlib import Path
from textwrap import dedent

import nbformat as nbf


PROJECT_ROOT = Path(__file__).resolve().parents[1]
NOTEBOOK_PATH = PROJECT_ROOT / "notebooks" / "01_eda_data_profiling.ipynb"


def md(source: str) -> nbf.NotebookNode:
    return nbf.v4.new_markdown_cell(dedent(source).strip() + "\n")


def code(source: str) -> nbf.NotebookNode:
    return nbf.v4.new_code_cell(dedent(source).strip() + "\n")


def build_notebook() -> nbf.NotebookNode:
    cells = [
        md(
            """
            # ISAP EDA & Data Profiling Notebook

            Notebook นี้เป็นหลักฐานการสำรวจข้อมูลจริงของทั้ง 2 datasets คือ OCSC และ CGD ก่อนนำไปออกแบบ data warehouse และ pipeline

            หลักการใช้งาน:

            - ใช้ notebook เพื่ออธิบาย EDA/profiling ให้อ่านตามง่าย
            - ใช้ production-style code ใน `src/isap_pipeline/` สำหรับ pipeline จริง
            - ใช้รายงานใน `docs/` เป็น final answer สำหรับส่งผู้ตรวจ
            """
        ),
        md(
            """
            ## สรุปสำคัญก่อนอ่าน

            ส่วนนี้คือบทสรุปสั้น ๆ สำหรับผู้อ่านที่ต้องการเห็นผลสำคัญก่อนลงรายละเอียด

            - OCSC workbook มี 68 sheets และเป็น Excel แบบรายงาน มี cover/index, merged cells, formula cells, multi-row headers, subtotal/total rows และ wide tables หลายรูปแบบ
            - CGD workbook มี 15 sheets โครงสร้างสม่ำเสมอกว่า แต่ยังมีหัวตาราง 2 ชั้น และต้องแยกความหมายระหว่าง `เบิกจ่าย` กับ `ใช้จ่าย`
            - แนวทางที่เหมาะคือเก็บ raw cells ทุก sheet ก่อน แล้ว normalize เฉพาะ sheet/table ที่มี grain ชัดเจนเข้าสู่ staging/mart
            - EDA พบและแก้ปัญหาเพิ่ม 2 จุด: OCSC มีแถว `ร้อยละ` ที่ไม่ควรถูกนับเป็น headcount และ CGD มีแถว `รวม` อยู่คนละคอลัมน์กับชื่อ entity ปกติ
            - การเชื่อม OCSC กับ CGD ทำได้ระดับ demo ด้วย normalized Thai entity name แต่ production ควรมี master agency mapping
            """
        ),
        md(
            """
            ## 1. Setup

            ส่วนนี้โหลด dependency และ helper function จาก package จริง เพื่อให้ notebook ไม่แยก logic ออกจาก pipeline
            """
        ),
        code(
            r"""
            from pathlib import Path
            import sys
            import warnings

            import matplotlib.pyplot as plt
            import pandas as pd
            from IPython.display import display
            from matplotlib import font_manager
            from openpyxl import load_workbook

            PROJECT_ROOT = Path.cwd()
            if PROJECT_ROOT.name == "notebooks":
                PROJECT_ROOT = PROJECT_ROOT.parent

            sys.path.insert(0, str(PROJECT_ROOT / "src"))

            from isap_pipeline.dq import run_data_quality_checks
            from isap_pipeline.excel_inspector import inspect_workbook
            from isap_pipeline.extract_cgd import extract_cgd_workbook
            from isap_pipeline.extract_ocsc import extract_ocsc_workbook
            from isap_pipeline.metadata import SourceFileMetadata, sha256_file

            warnings.filterwarnings("ignore", category=UserWarning, module=r"openpyxl\..*")
            pd.set_option("display.max_columns", 40)
            pd.set_option("display.max_colwidth", 90)

            available_fonts = {font.name for font in font_manager.fontManager.ttflist}
            thai_font_candidates = [
                "Noto Sans Thai Looped",
                "Leelawadee UI",
                "Tahoma",
                "Arial Unicode MS",
            ]
            THAI_FONT = next(
                (font for font in thai_font_candidates if font in available_fonts),
                None,
            )
            if THAI_FONT:
                plt.rcParams["font.family"] = THAI_FONT
            plt.rcParams["axes.unicode_minus"] = False

            OCSC_PATH = PROJECT_ROOT / "datasets" / "ocsc" / "thai-gov-manpower-2567.4.xlsx"
            CGD_PATH = PROJECT_ROOT / "datasets" / "cgd" / "2026.07.03.xlsx"

            assert OCSC_PATH.exists(), OCSC_PATH
            assert CGD_PATH.exists(), CGD_PATH
            """
        ),
        md(
            """
            ## 2. Source Metadata

            ตารางนี้ยืนยันตำแหน่งไฟล์, ขนาดไฟล์ และ SHA-256 hash สำหรับ lineage/idempotency
            """
        ),
        code(
            r"""
            source_files = pd.DataFrame(
                [
                    {
                        "dataset": "ocsc_government_manpower",
                        "path": str(OCSC_PATH.relative_to(PROJECT_ROOT)),
                        "sha256": sha256_file(OCSC_PATH),
                        "size_mb": round(OCSC_PATH.stat().st_size / 1024 / 1024, 2),
                    },
                    {
                        "dataset": "cgd_budget_execution",
                        "path": str(CGD_PATH.relative_to(PROJECT_ROOT)),
                        "sha256": sha256_file(CGD_PATH),
                        "size_mb": round(CGD_PATH.stat().st_size / 1024 / 1024, 2),
                    },
                ]
            )
            source_files
            """
        ),
        md(
            """
            ## 3. Workbook Inventory

            ใช้ `inspect_workbook()` เพื่อสรุป sheet count, cell count, merged cells, formula cells, blank rows และ guessed header row จากไฟล์จริง
            """
        ),
        code(
            r"""
            ocsc_profile = inspect_workbook(OCSC_PATH)
            cgd_profile = inspect_workbook(CGD_PATH)

            ocsc_sheets = pd.DataFrame(ocsc_profile["sheets"])
            cgd_sheets = pd.DataFrame(cgd_profile["sheets"])

            workbook_summary = pd.DataFrame(
                [
                    {
                        "dataset": "OCSC",
                        "sheet_count": ocsc_profile["sheet_count"],
                        "total_non_empty_cells": int(ocsc_sheets["non_empty_cells"].sum()),
                        "merged_cells": int(ocsc_sheets["merged_cell_count"].sum()),
                        "formula_cells": int(ocsc_sheets["formula_cell_count"].sum()),
                        "blank_rows": int(ocsc_sheets["blank_row_count"].sum()),
                    },
                    {
                        "dataset": "CGD",
                        "sheet_count": cgd_profile["sheet_count"],
                        "total_non_empty_cells": int(cgd_sheets["non_empty_cells"].sum()),
                        "merged_cells": int(cgd_sheets["merged_cell_count"].sum()),
                        "formula_cells": int(cgd_sheets["formula_cell_count"].sum()),
                        "blank_rows": int(cgd_sheets["blank_row_count"].sum()),
                    },
                ]
            )
            workbook_summary
            """
        ),
        code(
            r"""
            fig, axes = plt.subplots(1, 2, figsize=(11, 4))
            for ax, frame, title in [
                (axes[0], ocsc_sheets, "OCSC sheet types"),
                (axes[1], cgd_sheets, "CGD sheet types"),
            ]:
                counts = frame["sheet_type"].value_counts().sort_values(ascending=True)
                counts.plot(kind="barh", ax=ax)
                ax.set_title(title)
                ax.set_xlabel("sheet count")
                ax.set_ylabel("")
            plt.tight_layout()
            plt.show()
            """
        ),
        md(
            """
            ## 4. EDA Findings to Engineering Decisions

            ตารางนี้สรุปว่า observation จาก EDA ถูกแปลงเป็น design/pipeline decision อย่างไร เพื่อให้เห็นว่า notebook ไม่ได้เป็นแค่การดูไฟล์ Excel แต่เป็นหลักฐานของการออกแบบ data warehouse และ pipeline
            """
        ),
        code(
            r"""
            eda_decisions = pd.DataFrame(
                [
                    {
                        "finding": "workbook มี cover/index/report decoration",
                        "pipeline_decision": "เก็บ raw.workbook_sheets และ raw.cells เพื่อ audit ก่อนทำ staging",
                    },
                    {
                        "finding": "sheet มี header หลายแถวและ newline ในชื่อ column",
                        "pipeline_decision": "normalize/flatten header ใน clean.py และ extractor",
                    },
                    {
                        "finding": "CGD มีงบประจำ/งบลงทุน/รวม ในแนวกว้าง",
                        "pipeline_decision": "unpivot เป็น expense_category เพื่อให้ query ง่าย",
                    },
                    {
                        "finding": "CGD มีทั้งมุมเบิกจ่ายและใช้จ่าย",
                        "pipeline_decision": "แยกด้วย report_type เพื่อไม่ผสม metric คนละนิยาม",
                    },
                    {
                        "finding": "OCSC มีหลายมิติกำลังพลใน workbook เดียว",
                        "pipeline_decision": "ทำ fact แบบ metric-driven ด้วย metric_name และ metric_group",
                    },
                    {
                        "finding": "OCSC-CGD ไม่มี agency code กลางครบทุก sheet",
                        "pipeline_decision": "demo join ด้วย normalized Thai name; production ควรมี master agency mapping",
                    },
                ]
            )
            display(eda_decisions)

            cross_dataset_notes = pd.DataFrame(
                [
                    {
                        "topic": "agency/ministry name",
                        "risk": "join แล้วหลุดหรือจับคู่ผิด",
                        "current_handling": "normalize whitespace และ demo exact-name join",
                    },
                    {
                        "topic": "time grain",
                        "risk": "OCSC เป็น snapshot รายปี แต่ CGD เป็น as-of date",
                        "current_handling": "เก็บ fiscal_year, fiscal_year_be และ as_of_date แยกกัน",
                    },
                    {
                        "topic": "unit",
                        "risk": "headcount เทียบกับ million baht โดยตรงไม่ได้",
                        "current_handling": "แยก fact table และระบุ column unit ชัดเจน",
                    },
                    {
                        "topic": "total/subtotal rows",
                        "risk": "เกิด double count ใน query",
                        "current_handling": "tag entity_type และให้ analyst filter ได้",
                    },
                ]
            )
            display(cross_dataset_notes)
            """
        ),
        md(
            """
            ## 5. OCSC EDA

            OCSC เป็น workbook ที่ซับซ้อนกว่า เพราะมีหลาย report sections ในไฟล์เดียว ตารางด้านล่างแสดง sheet ที่ใหญ่ที่สุดและ sheet ที่มี merged/formula cells สูง
            """
        ),
        code(
            r"""
            profile_cols = [
                "sheet_name",
                "max_row",
                "max_column",
                "non_empty_cells",
                "merged_cell_count",
                "formula_cell_count",
                "blank_row_count",
                "guessed_header_row",
                "sheet_type",
            ]

            print("OCSC: top 12 sheets by non-empty cells")
            display(ocsc_sheets.sort_values("non_empty_cells", ascending=False)[profile_cols].head(12))

            print("OCSC: top 12 sheets by merged cells")
            display(ocsc_sheets.sort_values("merged_cell_count", ascending=False)[profile_cols].head(12))
            """
        ),
        code(
            r"""
            ocsc_issues = pd.DataFrame(
                [
                    {
                        "observed_issue": "มี cover/index/report decoration",
                        "why_it_matters": "ไม่ควร ingest ทุก sheet เป็น fact table เดียว",
                        "cleaning_strategy": "เก็บ raw cells ทุก sheet แล้วเลือก normalize เฉพาะ sheet ที่ grain ชัดเจน",
                    },
                    {
                        "observed_issue": "merged cells และ multi-row headers",
                        "why_it_matters": "header ไม่อยู่ใน row เดียวและชื่อ metric กระจายหลาย column",
                        "cleaning_strategy": "flatten header และ map metric เป็น long format",
                    },
                    {
                        "observed_issue": "subtotal/total ปะปนกับ detail rows",
                        "why_it_matters": "ถ้านับรวมกับ detail จะ double count",
                        "cleaning_strategy": "tag `entity_type` เป็น total/ministry/agency",
                    },
                    {
                        "observed_issue": "บาง cell เป็น formula/error token เช่น #REF!",
                        "why_it_matters": "numeric conversion ต้องไม่ crash และต้อง audit ได้",
                        "cleaning_strategy": "convert แบบ safe, เก็บ `source_value`, และใช้ DQ checks",
                    },
                ]
            )
            ocsc_issues
            """
        ),
        code(
            r"""
            def preview_sheet_by_name(path: Path, name_contains: str, rows: int = 12, cols: int = 12) -> pd.DataFrame:
                wb = load_workbook(path, data_only=True, read_only=True)
                ws = next(sheet for sheet in wb.worksheets if name_contains in sheet.title)
                values = []
                for row_idx in range(1, min(ws.max_row, rows) + 1):
                    values.append(
                        [
                            str(ws.cell(row_idx, col_idx).value or "").replace("\n", " ")[:80]
                            for col_idx in range(1, min(ws.max_column, cols) + 1)
                        ]
                    )
                wb.close()
                columns = [f"c{i}" for i in range(1, len(values[0]) + 1)]
                return pd.DataFrame(values, index=range(1, len(values) + 1), columns=columns)

            print("OCSC preview: sheet '17-29' workforce by agency")
            display(preview_sheet_by_name(OCSC_PATH, "17-29", rows=14, cols=12))

            print("OCSC preview: sheet '88-97' workforce profile")
            display(preview_sheet_by_name(OCSC_PATH, "88-97", rows=10, cols=20))
            """
        ),
        md(
            """
            ## 6. CGD EDA

            CGD โครงสร้างค่อนข้างสม่ำเสมอ แต่ต้องอ่านความหมายทางธุรกิจให้ถูก: `เบิกจ่าย` และ `ใช้จ่าย` ไม่ใช่ metric เดียวกัน
            """
        ),
        code(
            r"""
            print("CGD: all sheets")
            display(cgd_sheets[profile_cols].sort_values("sheet_name"))

            cgd_issues = pd.DataFrame(
                [
                    {
                        "observed_issue": "หัวตาราง 2 ชั้น",
                        "why_it_matters": "ชื่อ measure เช่น budget/disbursement/pct อยู่คนละ row กับ current/investment/total",
                        "cleaning_strategy": "หา group span จาก header row แล้ว unpivot เป็น long rows",
                    },
                    {
                        "observed_issue": "มีทั้ง sheet เบิกจ่ายและใช้จ่าย",
                        "why_it_matters": "เบิกจ่ายคือเงินจ่ายจริง ส่วนใช้จ่ายรวมมุมมอง commitment/PO บางส่วน",
                        "cleaning_strategy": "เก็บ `report_type` เป็น disbursement/expenditure",
                    },
                    {
                        "observed_issue": "หน่วยเงินเป็นล้านบาท",
                        "why_it_matters": "ต้องระบุ unit ชัดเจนเพื่อไม่ให้ analyst ตีความผิด",
                        "cleaning_strategy": "ใช้ column suffix `_million_baht`",
                    },
                    {
                        "observed_issue": "ชื่อหน่วยงานอาจซ้ำ แต่มี entity_code ต่างกัน",
                        "why_it_matters": "duplicate check ต้องใช้ grain ที่รวม code ไม่ใช่ใช้ชื่ออย่างเดียว",
                        "cleaning_strategy": "grain ของ CGD ใช้ source/sheet/entity/entity_code/report_type/expense_category",
                    },
                ]
            )
            cgd_issues
            """
        ),
        code(
            r"""
            print("CGD preview: summary sheet")
            display(preview_sheet_by_name(CGD_PATH, "สรุปภาพรวม", rows=12, cols=10))

            print("CGD preview: disbursement by agency")
            display(preview_sheet_by_name(CGD_PATH, "3.", rows=12, cols=15))

            print("CGD preview: expenditure by agency")
            display(preview_sheet_by_name(CGD_PATH, "4.", rows=12, cols=18))
            """
        ),
        md(
            """
            ## 7. Extracted Rows and Transform Validation

            ส่วนนี้เรียก extractor จริงจาก pipeline เพื่อให้เห็นว่า EDA เชื่อมกับ implementation ไม่ใช่แค่การดู Excel ด้วยตา
            """
        ),
        code(
            r"""
            run_id = "notebook-eda"
            ocsc_meta = SourceFileMetadata(
                dataset_name="ocsc_government_manpower",
                source_name="OCSC government workforce statistics",
                path=OCSC_PATH,
                sha256=sha256_file(OCSC_PATH),
                fiscal_year=2024,
                fiscal_year_be=2567,
            )
            cgd_meta = SourceFileMetadata(
                dataset_name="cgd_budget_execution",
                source_name="CGD budget disbursement statistics",
                path=CGD_PATH,
                sha256=sha256_file(CGD_PATH),
                fiscal_year=2026,
                fiscal_year_be=2569,
            )

            ocsc_extract = extract_ocsc_workbook(OCSC_PATH, ocsc_meta, run_id)
            cgd_extract = extract_cgd_workbook(CGD_PATH, cgd_meta, run_id)
            ocsc_workforce = pd.concat(
                [ocsc_extract.workforce_agency, ocsc_extract.workforce_profile],
                ignore_index=True,
            )

            extracted_summary = pd.DataFrame(
                [
                    {
                        "dataset": "OCSC",
                        "raw_cell_rows": len(ocsc_extract.raw_cells),
                        "normalized_rows": len(ocsc_workforce),
                        "target_table": "staging.ocsc_workforce",
                    },
                    {
                        "dataset": "CGD",
                        "raw_cell_rows": len(cgd_extract.raw_cells),
                        "normalized_rows": len(cgd_extract.budget_execution),
                        "target_table": "staging.cgd_budget_execution",
                    },
                ]
            )
            extracted_summary
            """
        ),
        code(
            r"""
            print("OCSC extracted rows by metric group")
            display(
                ocsc_workforce.groupby(["entity_type", "metric_group"], dropna=False)
                .size()
                .reset_index(name="rows")
                .sort_values("rows", ascending=False)
                .head(20)
            )

            print("CGD extracted rows by report type, entity type, and expense category")
            display(
                cgd_extract.budget_execution.groupby(
                    ["report_type", "entity_type", "expense_category"], dropna=False
                )
                .size()
                .reset_index(name="rows")
                .sort_values(["report_type", "entity_type", "expense_category"])
            )
            """
        ),
        code(
            r"""
            fig, axes = plt.subplots(1, 2, figsize=(12, 4))

            ocsc_workforce["metric_group"].value_counts().sort_values().plot(kind="barh", ax=axes[0])
            axes[0].set_title("OCSC normalized rows by metric group")
            axes[0].set_xlabel("rows")
            axes[0].set_ylabel("")

            cgd_extract.budget_execution["report_type"].value_counts().sort_values().plot(kind="barh", ax=axes[1])
            axes[1].set_title("CGD normalized rows by report type")
            axes[1].set_xlabel("rows")
            axes[1].set_ylabel("")

            plt.tight_layout()
            plt.show()
            """
        ),
        md(
            """
            ## 8. Content-Level Insights

            ส่วนนี้เพิ่มการวิเคราะห์เนื้อหาหลัง normalize โดยรักษา grain ให้ชัดเจน ไม่รวม total กับ detail และไม่เปรียบเทียบ OCSC ปี 2567 กับ CGD ปี 2569 เสมือนเป็นช่วงเวลาเดียวกัน
            """
        ),
        code(
            r"""
            ocsc_agency_headcount = ocsc_workforce[
                (ocsc_workforce["metric_name"] == "civil_servant")
                & (ocsc_workforce["entity_type"] == "agency")
                & ocsc_workforce["headcount"].notna()
            ].copy()
            ocsc_agency_headcount["entity_label"] = ocsc_agency_headcount.apply(
                lambda row: (
                    f"{row['ministry_name']} | {row['agency_name']}"
                    if pd.notna(row["ministry_name"])
                    else str(row["agency_name"])
                ),
                axis=1,
            )
            ocsc_top_agencies = ocsc_agency_headcount.nlargest(10, "headcount").sort_values(
                "headcount"
            )
            ocsc_top_agencies["chart_label"] = (
                ocsc_top_agencies["entity_label"]
                if THAI_FONT
                else [f"Agency {index + 1}" for index in range(len(ocsc_top_agencies))]
            )

            fig, ax = plt.subplots(figsize=(10, 5.5))
            ax.barh(
                ocsc_top_agencies["chart_label"],
                ocsc_top_agencies["headcount"],
                color="#2F6B9A",
                edgecolor="#1F2933",
                linewidth=0.6,
            )
            ax.set_title("OCSC: top 10 agency rows by civil-servant headcount")
            ax.set_xlabel("people, fiscal year 2567 (BE)")
            ax.set_ylabel("")
            ax.grid(axis="x", color="#D9DEE3", linewidth=0.6)
            ax.set_axisbelow(True)
            for container in ax.containers:
                ax.bar_label(container, fmt="{:,.0f}", padding=3, fontsize=8)
            plt.tight_layout()
            plt.show()

            display(
                ocsc_top_agencies[["chart_label", "ministry_name", "agency_name", "headcount"]]
                .sort_values("headcount", ascending=False)
                .reset_index(drop=True)
            )
            """
        ),
        md(
            """
            กราฟ OCSC ใช้เฉพาะ `entity_type='agency'` และ `metric_name='civil_servant'` พร้อมแสดงกระทรวงใน label เพื่อแยกชื่อทั่วไปที่ซ้ำกัน เช่น สำนักงานปลัดกระทรวง กราฟนี้เป็น ranking ตาม source row ไม่ใช่การรวมทุก employment type
            """
        ),
        code(
            r"""
            cgd_disbursement = cgd_extract.budget_execution[
                (cgd_extract.budget_execution["report_type"] == "disbursement")
                & (cgd_extract.budget_execution["expense_category"] == "total")
                & (~cgd_extract.budget_execution["entity_type"].isin(["total", "summary"]))
                & cgd_extract.budget_execution["disbursement_pct"].between(0, 100)
            ].copy()

            distribution_summary = (
                cgd_disbursement.groupby("entity_type")["disbursement_pct"]
                .agg(observations="count", median_pct="median", mean_pct="mean", min_pct="min", max_pct="max")
                .round(2)
                .sort_values("observations", ascending=False)
            )
            display(distribution_summary)

            entity_order = distribution_summary.index.tolist()
            boxplot_values = [
                cgd_disbursement.loc[
                    cgd_disbursement["entity_type"] == entity_type, "disbursement_pct"
                ].to_numpy()
                for entity_type in entity_order
            ]
            entity_labels = [
                f"{entity_type} (n={len(values)})"
                for entity_type, values in zip(entity_order, boxplot_values, strict=True)
            ]

            fig, ax = plt.subplots(figsize=(10, 5))
            boxplot = ax.boxplot(
                boxplot_values,
                tick_labels=entity_labels,
                vert=False,
                patch_artist=True,
                medianprops={"color": "#1F2933", "linewidth": 1.4},
                boxprops={"edgecolor": "#2F6B9A", "linewidth": 1.0},
                whiskerprops={"color": "#657786"},
                capprops={"color": "#657786"},
                flierprops={"marker": "o", "markersize": 3, "markerfacecolor": "#D99032"},
            )
            for box in boxplot["boxes"]:
                box.set_facecolor("#DCEAF4")
            ax.set_title("CGD: distribution of total disbursement rate by entity type")
            ax.set_xlabel("disbursement rate (%), as of 3 July 2026")
            ax.set_xlim(0, 105)
            ax.grid(axis="x", color="#D9DEE3", linewidth=0.6)
            ax.set_axisbelow(True)
            plt.tight_layout()
            plt.show()
            """
        ),
        md(
            """
            Box plot แสดง median, spread และ outlier ของอัตราเบิกจ่ายโดยใช้แถวระดับ detail ที่อยู่ใน grain เดียวกัน (`report_type='disbursement'`, `expense_category='total'`) จำนวน observation ในแต่ละกลุ่มไม่เท่ากัน จึงแสดง `n` กำกับ และไม่ควรใช้กราฟนี้สรุปสาเหตุของความแตกต่าง
            """
        ),
        code(
            r"""
            def build_reconciliation(frame: pd.DataFrame, measure: str) -> pd.DataFrame:
                rows = []
                group_cols = ["sheet_name", "report_type", "expense_category"]
                for group_key, group in frame[frame[measure].notna()].groupby(group_cols):
                    published_rows = group[group["entity_type"] == "total"]
                    detail_rows = group[group["entity_type"] != "total"]
                    if len(published_rows) != 1 or detail_rows.empty:
                        continue
                    published_total = float(published_rows[measure].iloc[0])
                    detail_sum = float(detail_rows[measure].sum())
                    difference = detail_sum - published_total
                    tolerance = max(0.01, abs(published_total) * 1e-6)
                    rows.append(
                        {
                            "sheet_name": group_key[0],
                            "report_type": group_key[1],
                            "expense_category": group_key[2],
                            "detail_rows": len(detail_rows),
                            "published_total": published_total,
                            "detail_sum": detail_sum,
                            "difference": difference,
                            "within_tolerance": abs(difference) <= tolerance,
                        }
                    )
                return pd.DataFrame(rows)

            cgd_reconciliation = build_reconciliation(
                cgd_extract.budget_execution, "disbursement_million_baht"
            )
            print(
                "CGD reconciliation groups within tolerance:",
                f"{int(cgd_reconciliation['within_tolerance'].sum())}/{len(cgd_reconciliation)}",
            )
            display(
                cgd_reconciliation.assign(
                    published_total=lambda frame: frame["published_total"].round(4),
                    detail_sum=lambda frame: frame["detail_sum"].round(4),
                    difference=lambda frame: frame["difference"].round(6),
                ).sort_values(["sheet_name", "expense_category"])
            )

            def normalized_names(series: pd.Series) -> set[str]:
                return set(
                    series.dropna()
                    .astype(str)
                    .str.lower()
                    .str.replace(r"\s+", "", regex=True)
                    .loc[lambda values: values.ne("")]
                )

            ocsc_names = normalized_names(ocsc_workforce["agency_name"])
            cgd_names = normalized_names(cgd_extract.budget_execution["entity_name"])
            exact_matches = ocsc_names & cgd_names
            cross_source_coverage = pd.DataFrame(
                [
                    {
                        "source": "OCSC",
                        "distinct_normalized_names": len(ocsc_names),
                        "exact_matches": len(exact_matches),
                        "coverage_pct": round(100 * len(exact_matches) / len(ocsc_names), 1),
                    },
                    {
                        "source": "CGD",
                        "distinct_normalized_names": len(cgd_names),
                        "exact_matches": len(exact_matches),
                        "coverage_pct": round(100 * len(exact_matches) / len(cgd_names), 1),
                    },
                ]
            )
            display(cross_source_coverage)
            """
        ),
        md(
            """
            Reconciliation ใช้ยอด `รวม` ที่เผยแพร่ในแต่ละ CGD sheet เทียบกับผลรวม detail ภายใน tolerance สำหรับ floating-point rounding ส่วน exact-name coverage เป็นเพียง diagnostic ของ join candidate ไม่ใช่ accuracy เพราะสองแหล่งมีขอบเขต entity ต่างกันและยังไม่มี agency master ที่ผ่าน human review
            """
        ),
        md(
            """
            ## 9. Data Quality Checks

            DQ checks ใช้กฎที่ตรงกับ business grain เช่น CGD ต้องรวม `entity_code` ใน duplicate key เพราะบางหน่วยงานชื่อซ้ำได้ และตรวจยอดรวมกับ detail เมื่อ source มี total row ที่เปรียบเทียบได้
            """
        ),
        code(
            r"""
            dq_result = run_data_quality_checks(cgd_extract.budget_execution, ocsc_workforce, run_id)
            dq_result
            """
        ),
        md(
            """
            ## 10. Takeaways

            - ทั้ง 2 datasets ถูกแยก source, extractor, staging และ mart ชัดเจน
            - OCSC ต้องระวัง report workbook structure จึงใช้แนว raw-first และ normalize เฉพาะ grain ที่ชัดเจนก่อน
            - CGD เหมาะกับการ flatten/unpivot เพราะ header structure ค่อนข้างซ้ำกันหลาย sheet
            - Content-level EDA ต้องกำหนด metric, entity level และ total/detail filter ก่อนทำกราฟทุกครั้ง
            - CGD total rows ที่กู้คืนจากคอลัมน์ A ทำให้ตรวจ reconciliation กับ detail ได้จริง
            - การ join ระหว่าง OCSC และ CGD ยังควรถือเป็น demo จนกว่าจะมี master agency mapping
            - Notebook นี้ควรใช้เป็น EDA evidence ส่วนการ demo pipeline ให้ใช้ CLI ใน README
            """
        ),
    ]

    return nbf.v4.new_notebook(
        cells=cells,
        metadata={
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {
                "name": "python",
                "pygments_lexer": "ipython3",
            },
        },
    )


def main() -> None:
    NOTEBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    nbf.write(build_notebook(), NOTEBOOK_PATH)
    print(f"Wrote {NOTEBOOK_PATH}")


if __name__ == "__main__":
    main()
