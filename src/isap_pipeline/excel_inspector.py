from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from isap_pipeline.clean import normalize_text

@dataclass(frozen=True)
class SheetProfile:
    sheet_name: str
    max_row: int
    max_column: int
    non_empty_cells: int
    merged_cell_count: int
    formula_cell_count: int
    blank_row_count: int
    blank_column_count: int
    guessed_header_row: int | None
    sheet_type: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "max_row": self.max_row,
            "max_column": self.max_column,
            "non_empty_cells": self.non_empty_cells,
            "merged_cell_count": self.merged_cell_count,
            "formula_cell_count": self.formula_cell_count,
            "blank_row_count": self.blank_row_count,
            "blank_column_count": self.blank_column_count,
            "guessed_header_row": self.guessed_header_row,
            "sheet_type": self.sheet_type,
        }

def inspect_workbook(path: str | Path) -> dict[str, Any]:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    profiles = [profile_sheet(ws).as_dict() for ws in workbook.worksheets]
    workbook.close()
    return {
        "filename": workbook_path.name,
        "path": str(workbook_path),
        "sheet_count": len(profiles),
        "sheets": profiles,
    }

def profile_sheet(ws: Worksheet) -> SheetProfile:
    non_empty = 0
    formula_count = 0
    row_non_empty: dict[int, int] = {}
    col_non_empty: dict[int, int] = {}
    text_sample: list[str] = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            non_empty += 1
            row_non_empty[cell.row] = row_non_empty.get(cell.row, 0) + 1
            col_non_empty[cell.column] = col_non_empty.get(cell.column, 0) + 1
            if cell.data_type == "f":
                formula_count += 1
            if len(text_sample) < 80:
                text_sample.append(normalize_text(cell.value))

    blank_rows = max(ws.max_row - len(row_non_empty), 0)
    blank_cols = max(ws.max_column - len(col_non_empty), 0)
    guessed_header = guess_header_row(ws)
    return SheetProfile(
        sheet_name=ws.title,
        max_row=ws.max_row,
        max_column=ws.max_column,
        non_empty_cells=non_empty,
        merged_cell_count=len(ws.merged_cells.ranges),
        formula_cell_count=formula_count,
        blank_row_count=blank_rows,
        blank_column_count=blank_cols,
        guessed_header_row=guessed_header,
        sheet_type=classify_sheet(ws.title, " ".join(text_sample), non_empty, ws.max_row, ws.max_column),
    )

def guess_header_row(ws: Worksheet) -> int | None:
    candidates: list[tuple[int, int, int]] = []
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        texts = [normalize_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        non_empty = sum(1 for item in texts if item)
        keyword_score = sum(
            1
            for item in texts
            if any(keyword in item for keyword in ["รายการ", "ส่วนราชการ", "หน่วยงาน", "กระทรวง", "จังหวัด", "รวม"])
        )
        if non_empty:
            candidates.append((keyword_score, non_empty, row_idx))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][2]

def classify_sheet(title: str, text_sample: str, non_empty: int, max_row: int, max_col: int) -> str:
    text = f"{title} {text_sample}"
    if non_empty <= 2:
        return "cover_or_blank"
    if "สารบัญ" in text:
        return "index"
    if "สรุปภาพรวม" in text or "ภาพรวม" in text:
        return "summary"
    if "จังหวัด" in text:
        return "geography"
    if any(keyword in text for keyword in ["กระทรวง", "หน่วยงาน", "ส่วนราชการ"]):
        return "agency"
    if any(keyword in text for keyword in ["เบิกจ่าย", "ใช้จ่าย", "งบประมาณ"]):
        return "budget_execution"
    if any(keyword in text for keyword in ["ช่วงอายุ", "เพศ", "ระดับการศึกษา"]):
        return "workforce_profile"
    if max_col > 20:
        return "wide_report"
    if max_row > 200:
        return "detail"
    return "report"

def workbook_sheet_inventory(path: str | Path) -> list[dict[str, Any]]:
    return inspect_workbook(path)["sheets"]