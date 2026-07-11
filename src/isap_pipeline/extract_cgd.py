from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from isap_pipeline.clean import canonical_entity_name, is_total_label, parse_thai_date, to_number
from isap_pipeline.metadata import SourceFileMetadata

@dataclass(frozen=True)
class CgdExtract:
    budget_execution: pd.DataFrame
    raw_cells: pd.DataFrame
    as_of_date: date | None

def extract_cgd_workbook(path: str | Path, meta: SourceFileMetadata, run_id: str) -> CgdExtract:
    workbook = load_workbook(path, data_only=True, read_only=False)
    raw_frames = []
    budget_rows: list[dict[str, Any]] = []
    as_of_date: date | None = None

    for sheet_index, ws in enumerate(workbook.worksheets, start=1):
        raw_frames.append(_raw_cells(ws, meta, run_id, sheet_index))
        sheet_date = _extract_as_of_date(ws)
        as_of_date = as_of_date or sheet_date
        if _is_summary_sheet(ws):
            budget_rows.extend(_parse_summary_sheet(ws, meta, run_id, sheet_date))
        else:
            budget_rows.extend(_parse_detail_sheet(ws, meta, run_id, sheet_date))

    workbook.close()
    raw_cells = pd.concat(raw_frames, ignore_index=True) if raw_frames else pd.DataFrame()
    budget_execution = pd.DataFrame(budget_rows)
    return CgdExtract(budget_execution=budget_execution, raw_cells=raw_cells, as_of_date=as_of_date)

def _raw_cells(ws: Worksheet, meta: SourceFileMetadata, run_id: str, sheet_index: int) -> pd.DataFrame:
    rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            rows.append(
                {
                    "ingestion_run_id": run_id,
                    "dataset_name": meta.dataset_name,
                    "source_file_hash": meta.sha256,
                    "sheet_index": sheet_index,
                    "sheet_name": ws.title,
                    "row_number": cell.row,
                    "column_number": cell.column,
                    "cell_value": str(cell.value),
                }
            )
    return pd.DataFrame(rows)

def _extract_as_of_date(ws: Worksheet) -> date | None:
    for row_idx in range(1, min(ws.max_row, 24) + 1):
        for col_idx in range(1, min(ws.max_column, 4) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value and "วันที่" in str(value):
                parsed = parse_thai_date(str(value))
                if parsed:
                    return parsed
    return None

def _is_summary_sheet(ws: Worksheet) -> bool:
    title = str(ws.title)
    first_header = str(ws.cell(4, 1).value or "")
    return "สรุปภาพรวม" in title or first_header == "รายการ"

def _parse_summary_sheet(
    ws: Worksheet, meta: SourceFileMetadata, run_id: str, as_of_date: date | None
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(6, ws.max_row + 1):
        item_name = canonical_entity_name(ws.cell(row_idx, 1).value)
        if not item_name or item_name.startswith("หมายเหตุ") or item_name.startswith("ที่มา"):
            break
        rows.append(
            _base_row(meta, run_id, ws.title, row_idx, "summary", item_name, None, "total", as_of_date)
            | {
                "budget_after_transfer_million_baht": to_number(ws.cell(row_idx, 2).value),
                "allocated_million_baht": to_number(ws.cell(row_idx, 3).value),
                "po_reserved_debt_million_baht": None,
                "disbursement_million_baht": to_number(ws.cell(row_idx, 6).value),
                "disbursement_pct": to_number(ws.cell(row_idx, 7).value),
                "expenditure_million_baht": to_number(ws.cell(row_idx, 4).value),
                "expenditure_pct": to_number(ws.cell(row_idx, 5).value),
                "monthly_target_gap_pct": None,
                "remaining_million_baht": to_number(ws.cell(row_idx, 8).value),
                "remaining_pct": to_number(ws.cell(row_idx, 9).value),
            }
        )
    return rows

def _parse_detail_sheet(
    ws: Worksheet, meta: SourceFileMetadata, run_id: str, as_of_date: date | None
) -> list[dict[str, Any]]:
    header_row = 4
    subheader_row = 5
    entity_col = _find_entity_column(ws, header_row)
    code_col = _find_code_column(ws, header_row, subheader_row)
    entity_type = _entity_type_from_sheet(ws.title)
    group_spans = _find_group_spans(ws, header_row, code_col)
    rows: list[dict[str, Any]] = []

    for row_idx in range(6, ws.max_row + 1):
        entity_name = canonical_entity_name(ws.cell(row_idx, entity_col).value)
        first_column_label = canonical_entity_name(ws.cell(row_idx, 1).value)
        if not entity_name and is_total_label(first_column_label):
            entity_name = first_column_label
        if not entity_name or entity_name.startswith("หมายเหตุ") or entity_name.startswith("ที่มา"):
            continue
        if is_total_label(entity_name):
            entity_type_current = "total"
        else:
            entity_type_current = entity_type
        entity_code = str(ws.cell(row_idx, code_col).value).strip() if code_col and ws.cell(row_idx, code_col).value else None
        for expense_category, start_col, end_col in group_spans:
            measures = _measure_values(ws, row_idx, subheader_row, start_col, end_col)
            rows.append(
                _base_row(
                    meta,
                    run_id,
                    ws.title,
                    row_idx,
                    entity_type_current,
                    entity_name,
                    entity_code,
                    expense_category,
                    as_of_date,
                )
                | measures
            )
    return rows

def _base_row(
    meta: SourceFileMetadata,
    run_id: str,
    sheet_name: str,
    row_number: int,
    entity_type: str,
    entity_name: str,
    entity_code: str | None,
    expense_category: str,
    as_of_date: date | None,
) -> dict[str, Any]:
    return {
        "ingestion_run_id": run_id,
        "dataset_name": meta.dataset_name,
        "source_file_hash": meta.sha256,
        "sheet_name": sheet_name,
        "row_number": row_number,
        "fiscal_year": meta.fiscal_year,
        "fiscal_year_be": meta.fiscal_year_be,
        "as_of_date": as_of_date.isoformat() if as_of_date else meta.as_of_date,
        "report_type": _report_type_from_sheet(sheet_name),
        "entity_type": entity_type,
        "entity_name": entity_name,
        "entity_code": entity_code,
        "expense_category": expense_category,
    }


def _report_type_from_sheet(sheet_name: str) -> str:
    normalized = sheet_name.strip()
    if "ใช้จ่าย" in normalized or normalized.endswith("(ใช้"):
        return "expenditure"
    return "disbursement"

def _entity_type_from_sheet(sheet_name: str) -> str:
    if "กระทรวง" in sheet_name:
        return "ministry"
    if "จังหวัด" in sheet_name:
        return "province"
    if "เทศบาล" in sheet_name:
        return "municipality"
    if "อบจ" in sheet_name:
        return "provincial_admin_org"
    if "รัฐวิสาหกิจ" in sheet_name:
        return "state_enterprise"
    if "กองทุน" in sheet_name:
        return "fund"
    return "agency"

def _find_entity_column(ws: Worksheet, header_row: int) -> int:
    for col_idx in range(1, ws.max_column + 1):
        value = str(ws.cell(header_row, col_idx).value or "")
        if any(keyword in value for keyword in ["กระทรวง", "หน่วยงาน", "จังหวัด", "เทศบาล", "รัฐวิสาหกิจ", "กองทุน"]):
            return col_idx
    return 2

def _find_code_column(ws: Worksheet, header_row: int, subheader_row: int) -> int | None:
    for col_idx in range(1, ws.max_column + 1):
        value = f"{ws.cell(header_row, col_idx).value or ''} {ws.cell(subheader_row, col_idx).value or ''}"
        if "รหัส" in value:
            return col_idx
    return None

def _find_group_spans(ws: Worksheet, header_row: int, code_col: int | None) -> list[tuple[str, int, int]]:
    starts: list[tuple[str, int]] = []
    for col_idx in range(1, ws.max_column + 1):
        label = str(ws.cell(header_row, col_idx).value or "").strip()
        if label in {"รายจ่ายประจำ", "รายจ่ายลงทุน", "รวม"}:
            starts.append((_expense_category(label), col_idx))
    spans: list[tuple[str, int, int]] = []
    upper_bound = (code_col - 1) if code_col else ws.max_column
    for idx, (label, start_col) in enumerate(starts):
        next_start = starts[idx + 1][1] if idx + 1 < len(starts) else upper_bound + 1
        spans.append((label, start_col, min(next_start - 1, upper_bound)))
    return spans

def _expense_category(label: str) -> str:
    if label == "รายจ่ายประจำ":
        return "current"
    if label == "รายจ่ายลงทุน":
        return "investment"
    return "total"

def _measure_values(ws: Worksheet, row_idx: int, subheader_row: int, start_col: int, end_col: int) -> dict[str, float | None]:
    result: dict[str, float | None] = {
        "budget_after_transfer_million_baht": None,
        "allocated_million_baht": None,
        "po_reserved_debt_million_baht": None,
        "disbursement_million_baht": None,
        "disbursement_pct": None,
        "expenditure_million_baht": None,
        "expenditure_pct": None,
        "monthly_target_gap_pct": None,
        "remaining_million_baht": None,
        "remaining_pct": None,
    }
    for col_idx in range(start_col, end_col + 1):
        header = str(ws.cell(subheader_row, col_idx).value or "")
        value = to_number(ws.cell(row_idx, col_idx).value)
        if "วงเงิน" in header:
            result["budget_after_transfer_million_baht"] = value
        elif "จัดสรร" in header:
            result["allocated_million_baht"] = value
        elif "PO" in header or "สำรอง" in header:
            result["po_reserved_debt_million_baht"] = value
        elif "สูง/ต่ำ" in header or "เป้าหมาย" in header:
            result["monthly_target_gap_pct"] = value
        elif "ร้อยละ" in header and "เบิกจ่าย" in header:
            result["disbursement_pct"] = value
        elif "ร้อยละ" in header and "ใช้จ่าย" in header:
            result["expenditure_pct"] = value
        elif "เบิกจ่าย" in header:
            result["disbursement_million_baht"] = value
        elif "การใช้จ่าย" in header:
            result["expenditure_million_baht"] = value
    return result
