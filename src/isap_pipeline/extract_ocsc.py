from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from isap_pipeline.clean import canonical_entity_name, fiscal_year_from_filename, is_total_label, to_int, to_number
from isap_pipeline.metadata import SourceFileMetadata

@dataclass(frozen=True)
class OcscExtract:
    workforce_agency: pd.DataFrame
    workforce_profile: pd.DataFrame
    raw_cells: pd.DataFrame

WORKFORCE_METRIC_COLUMNS = {
    4: "civil_servant",
    5: "permanent_employee_in_budget",
    6: "permanent_employee_out_budget",
    7: "temporary_employee_personnel_budget",
    8: "temporary_employee_other_budget",
    9: "temporary_employee_out_budget",
    10: "government_employee",
    11: "hired_employee",
    12: "university_employee",
}

PROFILE_METRIC_COLUMNS = {
    2: "civil_servant_total",
    3: "age_lt_21",
    4: "age_21_25",
    5: "age_26_30",
    6: "age_31_35",
    7: "age_36_40",
    8: "age_41_45",
    9: "age_46_50",
    10: "age_51_55",
    11: "age_56_60",
    12: "age_gt_60",
    13: "average_age",
    14: "gender_male",
    15: "gender_female",
    16: "female_pct",
    17: "education_below_bachelor",
    18: "education_bachelor",
    19: "education_master",
    20: "education_doctorate",
}

def extract_ocsc_workbook(path: str | Path, meta: SourceFileMetadata, run_id: str) -> OcscExtract:
    workbook = load_workbook(path, data_only=True, read_only=False)
    raw_frames = []
    agency_rows: list[dict[str, Any]] = []
    profile_rows: list[dict[str, Any]] = []

    fiscal_year, fiscal_year_be = meta.fiscal_year, meta.fiscal_year_be
    if fiscal_year is None:
        fiscal_year, fiscal_year_be = fiscal_year_from_filename(Path(path).name)

    for sheet_index, ws in enumerate(workbook.worksheets, start=1):
        raw_frames.append(_raw_cells(ws, meta, run_id, sheet_index))
        if _looks_like_workforce_agency_sheet(ws):
            agency_rows.extend(_parse_workforce_agency_sheet(ws, meta, run_id, fiscal_year, fiscal_year_be))
        if _looks_like_profile_sheet(ws):
            profile_rows.extend(_parse_profile_sheet(ws, meta, run_id, fiscal_year, fiscal_year_be))

    workbook.close()
    raw_cells = pd.concat(raw_frames, ignore_index=True) if raw_frames else pd.DataFrame()
    return OcscExtract(
        workforce_agency=pd.DataFrame(agency_rows),
        workforce_profile=pd.DataFrame(profile_rows),
        raw_cells=raw_cells,
    )

def _raw_cells(ws: Worksheet, meta: SourceFileMetadata, run_id: str, sheet_index: int) -> pd.DataFrame:
    rows = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            rows.append(
                {
                    "ingestion_run_id": run_id,
                    "dataset_name": meta.dataset_name,
                    "source_file_hash": meta.sha256,
                    "sheet_index": sheet_index,
                    "sheet_name": ws.title,
                    "row_number": cell.row,
                    "column_number": cell.column,
                    "cell_value": str(cell.value),
                }
            )
    return pd.DataFrame(rows)

def _looks_like_workforce_agency_sheet(ws: Worksheet) -> bool:
    first_cell = str(ws.cell(1, 1).value or "")
    header = " ".join(str(ws.cell(row, col).value or "") for row in range(1, 6) for col in range(1, min(ws.max_column, 12) + 1))
    return "ข้าราชการ ลูกจ้างประจำ" in first_cell and "พนักงานราชการ" in header

def _looks_like_profile_sheet(ws: Worksheet) -> bool:
    header = " ".join(str(ws.cell(row, col).value or "") for row in range(1, 4) for col in range(1, min(ws.max_column, 20) + 1))
    return all(keyword in header for keyword in ["ช่วงอายุ", "เพศ", "ระดับการศึกษา"])

def _parse_workforce_agency_sheet(
    ws: Worksheet,
    meta: SourceFileMetadata,
    run_id: str,
    fiscal_year: int | None,
    fiscal_year_be: int | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_ministry: str | None = None

    for row_idx in range(6, ws.max_row + 1):
        col_a = canonical_entity_name(ws.cell(row_idx, 1).value)
        col_c = canonical_entity_name(ws.cell(row_idx, 3).value)
        if not col_a and not col_c:
            continue
        if is_total_label(col_a):
            entity_name = "รวม"
            entity_type = "total"
            ministry_name = None
        elif col_a and not col_c:
            entity_name = col_a
            entity_type = "ministry"
            ministry_name = col_a
            current_ministry = col_a
        else:
            entity_name = col_c or col_a
            entity_type = "agency"
            ministry_name = current_ministry

        for col_idx, metric_name in WORKFORCE_METRIC_COLUMNS.items():
            value = to_int(ws.cell(row_idx, col_idx).value)
            if value is None:
                continue
            rows.append(
                {
                    "ingestion_run_id": run_id,
                    "dataset_name": meta.dataset_name,
                    "source_file_hash": meta.sha256,
                    "sheet_name": ws.title,
                    "row_number": row_idx,
                    "fiscal_year": fiscal_year,
                    "fiscal_year_be": fiscal_year_be,
                    "entity_type": entity_type,
                    "ministry_name": ministry_name,
                    "agency_name": entity_name,
                    "metric_name": metric_name,
                    "metric_group": "employment_type",
                    "headcount": value,
                    "percentage": None,
                    "source_value": str(ws.cell(row_idx, col_idx).value),
                    "source_unit": "person",
                }
            )
    return rows

def _parse_profile_sheet(
    ws: Worksheet,
    meta: SourceFileMetadata,
    run_id: str,
    fiscal_year: int | None,
    fiscal_year_be: int | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    current_ministry: str | None = None

    for row_idx in range(4, ws.max_row + 1):
        entity_name = canonical_entity_name(ws.cell(row_idx, 1).value)
        if not entity_name:
            continue
        if entity_name == "ร้อยละ":
            continue
        total = to_int(ws.cell(row_idx, 2).value)
        entity_type = "ministry" if total and entity_name.startswith(("กระทรวง", "สำนักนายกรัฐมนตรี")) else "agency"
        if entity_type == "ministry":
            current_ministry = entity_name
        if is_total_label(entity_name):
            entity_type = "total"

        for col_idx, metric_name in PROFILE_METRIC_COLUMNS.items():
            raw_value = ws.cell(row_idx, col_idx).value
            number = to_number(raw_value)
            if number is None:
                continue
            is_percentage = metric_name.endswith("_pct")
            is_average = metric_name == "average_age"
            rows.append(
                {
                    "ingestion_run_id": run_id,
                    "dataset_name": meta.dataset_name,
                    "source_file_hash": meta.sha256,
                    "sheet_name": ws.title,
                    "row_number": row_idx,
                    "fiscal_year": fiscal_year,
                    "fiscal_year_be": fiscal_year_be,
                    "entity_type": entity_type,
                    "ministry_name": None if entity_type in {"total", "ministry"} else current_ministry,
                    "agency_name": entity_name,
                    "metric_name": metric_name,
                    "metric_group": _profile_metric_group(metric_name),
                    "headcount": None if is_percentage or is_average else int(round(number)),
                    "percentage": number if is_percentage else None,
                    "source_value": str(raw_value),
                    "source_unit": "pct" if is_percentage else ("year" if is_average else "person"),
                }
            )
    return rows

def _profile_metric_group(metric_name: str) -> str:
    if metric_name.startswith("age_") or metric_name == "average_age":
        return "age"
    if metric_name.startswith("gender_") or metric_name == "female_pct":
        return "gender"
    if metric_name.startswith("education_"):
        return "education_level"
    return "total"
